from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import Any

from ig_corpus.config_schema import AppConfig, TargetsConfig
from ig_corpus.export_excel import export_corpus_workbook
from ig_corpus.llm_schema import LLMDecision
from ig_corpus.storage import SQLiteStateStore


def _decision(*, eligible: bool) -> LLMDecision:
    payload: dict[str, Any] = {
        "eligible": eligible,
        "eligibility_reasons": ["ok" if eligible else "reject"],
        "language": {"is_english": True, "confidence": 0.9},
        "topic": {
            "is_bodyweight_calisthenics": bool(eligible),
            "confidence": 0.9 if eligible else 0.1,
            "topic_notes": "test",
        },
        "commercial": {"is_exclusively_commercial": False, "signals": []},
        "caption_quality": {"is_analyzable": True, "issues": []},
        "tags": {
            "genre": "training_log" if eligible else "other",
            "narrative_labels": ["consistency"] if eligible else [],
            "discourse_moves": ["advice"] if eligible else [],
            "neoliberal_signals": [],
        },
        "overall_confidence": 0.9,
    }
    return LLMDecision.model_validate(payload)


class TestExportExcel(unittest.TestCase):
    def test_exports_required_sheets(self) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except Exception as e:  # pragma: no cover
            raise AssertionError("openpyxl is required for this test") from e

        cfg = AppConfig(targets=TargetsConfig(final_n=1, pool_n=1, sampling_seed=123))

        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "corpus.xlsx"

            with SQLiteStateStore.open(":memory:") as store:
                store.create_run(
                    config_hash="hash1",
                    sampling_seed=cfg.targets.sampling_seed,
                    versions={"python": "3.x"},
                    run_id="run_test",
                    started_at="2025-12-01T00:00:00+00:00",
                )

                store.upsert_raw_post(
                    post_key="id:1",
                    url="https://example.com/p/1",
                    actor_source="apify/actor",
                    raw_item={"url": "https://example.com/p/1", "caption": "x" * 80, "hashtags": ["tag1"]},
                    fetched_at="2025-12-01T00:00:00+00:00",
                )
                store.record_llm_decision(
                    post_key="id:1",
                    url="https://example.com/p/1",
                    model="gpt-5-nano",
                    decision=_decision(eligible=True),
                    created_at="2025-12-01T00:00:01+00:00",
                )

                store.upsert_raw_post(
                    post_key="id:2",
                    url="https://example.com/p/2",
                    actor_source="apify/actor",
                    raw_item={"url": "https://example.com/p/2", "caption": "y" * 80, "hashtags": ["tag2"]},
                    fetched_at="2025-12-01T00:00:00+00:00",
                )
                store.record_llm_decision(
                    post_key="id:2",
                    url="https://example.com/p/2",
                    model="gpt-5-nano",
                    decision=_decision(eligible=False),
                    created_at="2025-12-01T00:00:02+00:00",
                )

                export_corpus_workbook(cfg, store, out_path, run_id="run_test")

                # Pool target is met, so a final sample should be persisted.
                row = store.conn.execute(
                    "SELECT COUNT(1) FROM final_sample_runs WHERE run_id = ?",
                    ("run_test",),
                ).fetchone()
                self.assertIsNotNone(row)
                self.assertEqual(int(row[0]), 1)

                row2 = store.conn.execute(
                    "SELECT COUNT(1) FROM final_samples WHERE run_id = ?",
                    ("run_test",),
                ).fetchone()
                self.assertIsNotNone(row2)
                self.assertEqual(int(row2[0]), 1)

            self.assertTrue(out_path.exists())

            wb = load_workbook(out_path)
            self.assertIn("final500", wb.sheetnames)
            self.assertIn("eligible_pool", wb.sheetnames)
            self.assertIn("rejected", wb.sheetnames)
            self.assertIn("run_metadata", wb.sheetnames)
            self.assertIn("tag_summary", wb.sheetnames)

            ws = wb["final500"]
            rows = list(ws.iter_rows(values_only=True))
            self.assertGreaterEqual(len(rows), 2)  # header + at least one row

            header = [str(v) for v in rows[0]]
            self.assertIn("url", header)
            url_idx = header.index("url")
            self.assertEqual(rows[1][url_idx], "https://example.com/p/1")


if __name__ == "__main__":
    unittest.main()
